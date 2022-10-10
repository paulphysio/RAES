from django.http import HttpResponseRedirect
from django.shortcuts import render
import pandas as pd
import openpyxl as op
import smtplib, ssl
from email.message import EmailMessage
from resapp.models import activity
from .models import emailSender
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
# Create your views here.
def sendMail(request, pk):
    port = 465
    email = 'emilyjohnson25099@gmail.com'
    password = "wlwzkvalbpwebxot"

    activity_detail = activity.objects.get(id=pk)
    sheet = activity_detail.file_uploaded
    df = pd.read_excel(sheet)
    wb = op.load_workbook(sheet)
    wa = wb.active
    if df.columns.get_loc('Email'):
        email_index = df.columns.get_loc('Email')
    elif df.columns.get_loc('EMAIL'):
        email_index = df.columns.get_loc('EMAIL')
    elif df.columns.get_loc('email'):
        email_index = df.columns.get_loc('email')
    reallist = []
    f_msg = []
    for i in range(1, wa.max_row+1):
        lister=[]
        for j in range(1, wa.max_column+1):
            cell_obj = wa.cell(row=i, column=j)
            y = wa.cell(row=1, column=j).value
            lister.append(y)
            if j==(wa.max_column):
                f_msg.append(lister)
    for i in range(2, wa.max_row+1):
        list=[]
        for j in range(1, wa.max_column+1):
            cell_obj = wa.cell(row=i, column=j)
            x = wa.cell(row=i, column=j).value
            list.append(x)
            if j==(wa.max_column):
                reallist.append(list)
                
        print("sending " + str(list) +" to "+ str(wa.cell(row=i, column=email_index+ 1).value ))
        headmsg = (' | '.join(map(str, lister)))
        body_message =(' | '.join(map(str, list)))
        receiver = str(wa.cell(row=i, column=email_index+ 1).value)
        context = ssl.create_default_context()
        server=smtplib.SMTP_SSL("smtp.gmail.com", port, context=context)
        server.login(email, password)
        em = MIMEMultipart("alternative")
        em['From'] = 'The Course Adviser'
        em['To'] = receiver
        em['Subject'] = "Harmattan semester results"
        
        text = body_message
        results = []
        for i in range(len(lister)):
            for j in range(len(list)):
                if i == j:
                    #print(str(lister[i]) + ":" + str(list[j]))
                    result = str(lister[i]) + " : " + str(list[j])+";"
                    results.append(result)
        name= ('\n '.join(map(str, results)))
                    
                    
            
        html = f"""\
            <h1>This message contains the results for the harmattan semester.</h1>
            <h3>{name}</h3>
            
            """
            
        real_result = name
        # part1 = MIMEText(text, "plain")
        part2 = MIMEText(html, "html")
        part3 = MIMEText(real_result, "plain")
        # em.attach(part1)
        em.attach(part2)
        #em.attach(part3)
        server.sendmail(email, receiver, em.as_string())
        print("message sent")
        if request.method == 'POST':

            sender = request.user
            receiver_email = receiver
            message_sent = text + str(html) + real_result

            emailSender(sender=sender, receiver_email = receiver_email, message_sent = message_sent).save()
    return HttpResponseRedirect("/home")
